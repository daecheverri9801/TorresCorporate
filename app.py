# # ================================================================
# # Streamlit App :  Geocodificar Direcciones + Pre-Clasificación
# # Torres Corporate · Ing. Daniel Arango
# # ================================================================
# import os, zipfile, shutil, time, warnings, io, tempfile, re
# from pathlib import Path
# import pandas as pd
# import geopandas as gpd
# from utils_gsheets import update_master
# from shapely.geometry import Point, Polygon, LineString, MultiPolygon
# import googlemaps
# import streamlit as st
# from PIL import Image
# from lxml import etree
# from openpyxl.utils import get_column_letter
# from openpyxl.styles import Protection
# warnings.filterwarnings("ignore", category=UserWarning)

# # --------------------- CONFIGURACIÓN ----------------------------
# API_KEY  = os.getenv("GOOGLE_MAPS_API_KEY", "AIzaSyCkZG7fbor17mhs3NLjaThcChO-Pav67gA")
# DATA_DIR = "data"
# CACHE_FILE = "cache_geocoding.csv"
# KMZ_FILES  = [os.path.join(DATA_DIR, f) for f in os.listdir(DATA_DIR) if f.lower().endswith(".kmz")]
# LOGO_PATH  = "assets/logo_torres.jpg"
# LOGO_WIDTH = 120
# PROT_PASS  = "TorresC25"           # ← contraseña de protección

# st.set_page_config(page_title="Torres Corporate · Geolocalización",
#                    page_icon="🚚", layout="wide")

# # ------------------------- ESTILOS ------------------------------
# st.markdown(
#     """
#     <style>
#     :root{--primary:#005bea;--secondary:#00c6fb;--bg:#f7faff;--text:#10243c;}
#     body, section.main > div{background:var(--bg);} *{font-family:'Inter',sans-serif;}
#     body, p, div, span, li{color:var(--text)!important;}
#     .titulo{font-size:2.3rem;font-weight:700;color:var(--primary);}
#     hr.modern{border:none;height:4px;background:linear-gradient(90deg,var(--primary) 0%,var(--secondary) 100%);
#               margin:-4px 0 28px 0;border-radius:3px;}
#     .stButton>button{background:var(--primary);color:#fff;border:none;border-radius:6px;padding:0.55rem 1.2rem;font-weight:600;}
#     .stDownloadButton>button{background:var(--secondary);color:#fff!important;border:none;border-radius:6px;padding:0.55rem 1.2rem;font-weight:600;}
#     div[data-testid="stFileUploaderDropzone"]{border:2px dashed var(--primary)!important;background:#eaf2ff!important;}
#     </style>
#     """,
#     unsafe_allow_html=True
# )

# # --------------------------- HEADER -----------------------------
# c1, c2 = st.columns([1,4])
# with c1:
#     if os.path.exists(LOGO_PATH):
#         st.image(Image.open(LOGO_PATH), width=LOGO_WIDTH)
# with c2:
#     st.markdown('<div class="titulo">🚚 Torres Corporate</div>', unsafe_allow_html=True)
# st.markdown('<hr class="modern">', unsafe_allow_html=True)

# # --------------------- CARGAR / GUARDAR CACHÉ -------------------
# def load_cache():
#     if os.path.exists(CACHE_FILE) and os.path.getsize(CACHE_FILE) > 0:
#         return pd.read_csv(CACHE_FILE)
#     return pd.DataFrame(columns=["full_address", "latitude", "longitude", "geocode_status"])
# cache_df = load_cache()

# # ------------------- CARGA DE POLÍGONOS DESDE KMZ ---------------
# @st.cache_resource(show_spinner=False)
# def load_polygons():
#     def parse_coords(txt):
#         return [(float(lon), float(lat)) for lon, lat, *_ in
#                 (c.split(',') for c in re.split(r'\s+', txt.strip()) if c)]

#     def geom(pm):
#         txt = pm.xpath('.//*[local-name()="Point"]//*[local-name()="coordinates"]/text()')
#         if txt and (c := parse_coords(txt[0])):
#             return Point(c[0])
#         txt = pm.xpath('.//*[local-name()="Polygon"]//*[local-name()="outerBoundaryIs"]'
#                        '//*[local-name()="LinearRing"]//*[local-name()="coordinates"]/text()')
#         if txt and (c := parse_coords(txt[0])) and len(c) >= 3:
#             return Polygon(c)
#         txt = pm.xpath('.//*[local-name()="LineString"]//*[local-name()="coordinates"]/text()')
#         if txt and (c := parse_coords(txt[0])) and len(c) >= 2:
#             return LineString(c)
#         polys = []
#         for nodo in pm.xpath('.//*[local-name()="MultiGeometry"]//*[local-name()="Polygon"]'):
#             t = nodo.xpath('.//*[local-name()="outerBoundaryIs"]//*[local-name()="LinearRing"]//*[local-name()="coordinates"]/text()')
#             if t and (c := parse_coords(t[0])) and len(c) >= 3:
#                 polys.append(Polygon(c))
#         return MultiPolygon(polys) if polys else None

#     def kmz_to_gdf(kmz_path):
#         zona = Path(kmz_path).stem
#         tmp  = tempfile.mkdtemp()
#         try:
#             with zipfile.ZipFile(kmz_path) as zf:
#                 zf.extractall(tmp)
#             kmls = list(Path(tmp).rglob("*.kml")) + list(Path(tmp).rglob("*.KML"))
#             feats = []
#             for fp in kmls:
#                 try:
#                     tree = etree.parse(str(fp))
#                 except Exception:
#                     continue
#                 for pm in tree.xpath('.//*[local-name()="Placemark"]'):
#                     g = geom(pm)
#                     if g:
#                         name = (pm.xpath('.//*[local-name()="name"]/text()') or ["Sin_nombre"])[0]
#                         feats.append({"geometry": g, "zona": zona, "subzona": name})
#             return gpd.GeoDataFrame(feats, crs="EPSG:4326") if feats else None
#         finally:
#             shutil.rmtree(tmp, ignore_errors=True)

#     frames = [kmz_to_gdf(f) for f in KMZ_FILES]
#     frames = [g for g in frames if g is not None and not g.empty]
#     return pd.concat(frames, ignore_index=True).to_crs(4326) if frames else \
#            gpd.GeoDataFrame(columns=["geometry", "zona", "subzona"], crs="EPSG:4326")

# # ----------------------- GEOCODIFICACIÓN ------------------------
# def build_address(r):
#     comps = [str(r.get(c, "")).strip()
#              for c in ["Buyer Address1", "Buyer Address1 Number", "Buyer City"]
#              if str(r.get(c, "")).strip() and str(r.get(c, "")).lower() != "nan"]
#     return ", ".join(comps) + ", Colombia"

# def geocode_enhanced(cli, address, retries=3):
#     for _ in range(retries):
#         try:
#             res = cli.geocode(address, region="co", language="es")
#             if res:
#                 loc = res[0]["geometry"]["location"]
#                 return loc["lat"], loc["lng"], "completa"
#         except Exception:
#             time.sleep(0.8)
#     try:
#         city = address.split(",")[-2] + ", Colombia"
#         res  = cli.geocode(city, region="co", language="es")
#         if res:
#             loc = res[0]["geometry"]["location"]
#             return loc["lat"], loc["lng"], "ciudad"
#     except Exception:
#         pass
#     return None, None, "fallida"

# def geocode_with_cache(cli, address):
#     hit = cache_df.loc[cache_df.full_address == address]
#     if not hit.empty:
#         r = hit.iloc[0]
#         return r.latitude, r.longitude, r.geocode_status, True
#     lat, lon, stat = geocode_enhanced(cli, address)
#     cache_df.loc[len(cache_df)] = [address, lat, lon, stat]
#     return lat, lon, stat, False

# # ---------------- PRE-CLASIFICACIÓN (con protección) ------------
# def crear_preclasificacion(df_geo) -> io.BytesIO:
#     import openpyxl
#     total = len(df_geo)
#     base_df = df_geo[["Country", "Tracking Number", "subzona"]].rename(columns={"subzona": "Zona"})
#     reg_df  = pd.DataFrame({"Consecutivo": range(1, total + 1),
#                             "Tracking Number": ["" for _ in range(total)],
#                             "Zona": ["" for _ in range(total)]})

#     buf = io.BytesIO()
#     with pd.ExcelWriter(buf, engine="openpyxl") as writer:
#         base_df.to_excel(writer, sheet_name="Base", index=False)
#         reg_df.to_excel(writer, sheet_name="Registro", index=False, startrow=2)

#         ws_reg  = writer.sheets["Registro"]
#         ws_base = writer.sheets["Base"]

#         # — Estilo cabecera —
#         for c in ws_reg[1]:
#             c.font = c.font.copy(sz=18, bold=True)

#         last_row = ws_reg.max_row
#         # Fórmulas Zona (col C) y cabeceras dinámicas
#         for r in range(4, last_row + 1):
#             ws_reg[f"C{r}"] = f'=IFERROR(VLOOKUP($B{r},Base!$B:$C,2,FALSE),"")'
#         ws_reg["A1"] = f'=COUNTA(B4:B{last_row})'
#         ws_reg["B1"] = f'=IFERROR(LOOKUP(2,1/(B4:B{last_row}<>""),C4:C{last_row}),"")'
#         ws_reg["C1"] = f'=IFERROR(LOOKUP(2,1/(B4:B{last_row}<>""),B4:B{last_row}),"")'

#         for col, w in zip("ABC", (12, 25, 18)):
#             ws_reg.column_dimensions[col].width = w

#         # — Protección: Col A y B editables; Col C bloqueada —
#         for row in ws_reg.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=3):
#             for cell in row:
#                 lock = cell.column_letter == "C"
#                 cell.protection = Protection(locked=lock)

#         ws_reg.protection.sheet = True
#         ws_reg.protection.set_password(PROT_PASS)  # ← contraseña
#         ws_base.sheet_state = "veryHidden"

#     buf.seek(0)
#     return buf

# # --------------- PROCESAR UN ARCHIVO SUBIDO ---------------------
# def procesar_archivo(uploaded_file):
#     df = pd.read_excel(uploaded_file).copy().reset_index(drop=True)
#     df["full_address"] = df.apply(build_address, axis=1)

#     cli = googlemaps.Client(API_KEY)
#     lat, lon, stat = [], [], []
#     prog = st.progress(0, text=f"Geocodificando: {uploaded_file.name}")
#     for i, addr in enumerate(df.full_address):
#         la, lo, stt, hit = geocode_with_cache(cli, addr)
#         lat.append(la); lon.append(lo); stat.append(stt)
#         prog.progress((i + 1) / len(df))
#     prog.empty()
#     df["latitude"], df["longitude"], df["geocode_status"] = lat, lon, stat
#     cache_df.to_csv(CACHE_FILE, index=False)

#     # — Unión espacial —
#     gdf_pol = load_polygons()
#     gdf_pts = gpd.GeoDataFrame(
#         df[df.latitude.notna() & df.longitude.notna()].reset_index().rename(columns={'index': 'idx'}),
#         geometry=[Point(xy) for xy in zip(df.longitude, df.latitude)], crs=4326
#     )
#     join = gpd.sjoin(gdf_pts, gdf_pol[["zona", "subzona", "geometry"]],
#                      how="left", predicate="within").drop_duplicates(subset='idx').set_index('idx')
#     df = df.join(join[["zona", "subzona"]])

#     # — Reglas adicionales —
#     buyer_lower = df["Buyer City"].astype(str).str.lower()
#     mask_other  = buyer_lower == "other"
#     mask_bq     = buyer_lower == "barranquilla"
#     mask_override = (~mask_other) & (~mask_bq)

#     df.loc[mask_other, ["zona", "subzona"]] = "Verificar Zona"
#     df.loc[mask_override & df["Buyer City"].notna(), "subzona"] = df.loc[mask_override, "Buyer City"]

#     df["zona"].fillna("Verificar Zona", inplace=True)
#     df["subzona"].fillna("Verificar Zona", inplace=True)

#     sin_zona = (df["zona"] == "Verificar Zona").sum()

#     # — Crear archivo final —
#     pre_buf      = crear_preclasificacion(df)
    
#     # — Actualizar hoja master en Google Sheets —
#     try:
#         update_master(df)          # usa GSHEETS_URL y las credenciales definidas
#     except Exception as e:
#         st.warning(f"No se pudo actualizar el master en Google Sheets: {e}")

#     nombre_final = f"{Path(uploaded_file.name).stem}_zonificado.xlsx"

#     return {"filename": nombre_final,
#             "buffer": pre_buf,
#             "stats": {"total": len(df),
#                       "con_zona": len(df) - sin_zona,
#                       "sin_zona": sin_zona}}

# # --------------------------- INTERFAZ ---------------------------
# st.header("🗺️ Geolocalización y Pre-Clasificación")
# st.caption("Cargue hasta 5 archivos Excel a la vez.")

# uploads = st.file_uploader("⬆️  Seleccione archivos",
#                            type=["xlsx", "xls"], accept_multiple_files=True)

# if uploads and len(uploads) > 5:
#     st.error("Solo se permiten hasta 5 archivos simultáneamente.")

# if uploads and st.button("🚀 Procesar Archivos", type="primary"):
#     with st.spinner("Procesando archivos…"):
#         resultados = [procesar_archivo(u) for u in uploads[:5]]

#         zip_buf = io.BytesIO()
#         with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
#             for r in resultados:
#                 zf.writestr(r["filename"], r["buffer"].getvalue())
#         zip_buf.seek(0)

#     st.success("¡Procesamiento completo!")

#     for r in resultados:
#         with st.expander(f"📄 {r['filename']}"):
#             st.write(f"Total: {r['stats']['total']}")
#             st.write(f"Con zona: {r['stats']['con_zona']}")
#             st.write(f"Sin zona: {r['stats']['sin_zona']}")

#     st.download_button("📥 Descargar ZIP con archivos procesados",
#                        data=zip_buf,
#                        file_name="Archivos_Zonificados.zip",
#                        mime="application/zip")
# elif not uploads:
#     st.info("Arrastre o seleccione hasta 5 archivos Excel para habilitar el botón.")

# ================================================================
# Streamlit App · Geocodificar Direcciones + Pre-Clasificación
# Torres Corporate · Ing. Daniel Arango  ·  Versión v1.0-final
# ================================================================
import os, zipfile, shutil, time, warnings, io, tempfile, re
from pathlib import Path
import pandas as pd
import geopandas as gpd
from utils_gsheets import update_master
from shapely.geometry import Point, Polygon, LineString, MultiPolygon
import googlemaps
import streamlit as st
from PIL import Image
from lxml import etree
from openpyxl.utils import get_column_letter
from openpyxl.styles import Protection
warnings.filterwarnings("ignore", category=UserWarning)

# --------------------- CONFIGURACIÓN ----------------------------
API_KEY   = os.getenv("GOOGLE_MAPS_API_KEY", "AIzaSyCkZG7fbor17mhs3NLjaThcChO-Pav67gA")
DATA_DIR  = "data"
CACHE_FILE = "cache_geocoding.csv"
KMZ_FILES = [os.path.join(DATA_DIR, f) for f in os.listdir(DATA_DIR) if f.lower().endswith(".kmz")]
LOGO_PATH = "assets/logo_torres.jpg"
LOGO_WIDTH = 120
PROT_PASS = "TorresC25"                 # ← contraseña de protección

st.set_page_config(page_title="Torres Corporate · Geolocalización",
                   page_icon="🚚", layout="wide")

# ------------------------- ESTILOS ------------------------------
st.markdown(
    """
    <style>
    :root{--primary:#005bea;--secondary:#00c6fb;--bg:#f7faff;--text:#10243c;}
    body, section.main > div{background:var(--bg);} *{font-family:'Inter',sans-serif;}
    body, p, div, span, li{color:var(--text)!important;}
    .titulo{font-size:2.3rem;font-weight:700;color:var(--primary);}
    hr.modern{border:none;height:4px;background:linear-gradient(90deg,var(--primary) 0%,var(--secondary) 100%);
              margin:-4px 0 28px 0;border-radius:3px;}
    .stButton>button{background:var(--primary);color:#fff;border:none;border-radius:6px;padding:0.55rem 1.2rem;font-weight:600;}
    .stDownloadButton>button{background:var(--secondary);color:#fff!important;border:none;border-radius:6px;padding:0.55rem 1.2rem;font-weight:600;}
    div[data-testid="stFileUploaderDropzone"]{border:2px dashed var(--primary)!important;background:#eaf2ff!important;}
    </style>
    """,
    unsafe_allow_html=True
)

# --------------------------- HEADER -----------------------------
c1, c2 = st.columns([1, 4])
with c1:
    if os.path.exists(LOGO_PATH):
        st.image(Image.open(LOGO_PATH), width=LOGO_WIDTH)
with c2:
    st.markdown('<div class="titulo">🚚 Torres Corporate</div>', unsafe_allow_html=True)
st.markdown('<hr class="modern">', unsafe_allow_html=True)

# --------------------- CARGAR / GUARDAR CACHÉ -------------------
def load_cache():
    if os.path.exists(CACHE_FILE) and os.path.getsize(CACHE_FILE) > 0:
        return pd.read_csv(CACHE_FILE)
    return pd.DataFrame(columns=["full_address", "latitude", "longitude", "geocode_status"])
cache_df = load_cache()

# ------------------- CARGA DE POLÍGONOS DESDE KMZ ---------------
@st.cache_resource(show_spinner=False)
def load_polygons():
    def parse_coords(txt):
        return [(float(lon), float(lat)) for lon, lat, *_ in
                (c.split(',') for c in re.split(r'\s+', txt.strip()) if c)]

    def geom(pm):
        txt = pm.xpath('.//*[local-name()="Point"]//*[local-name()="coordinates"]/text()')
        if txt and (c := parse_coords(txt[0])):
            return Point(c[0])
        txt = pm.xpath('.//*[local-name()="Polygon"]//*[local-name()="outerBoundaryIs"]'
                       '//*[local-name()="LinearRing"]//*[local-name()="coordinates"]/text()')
        if txt and (c := parse_coords(txt[0])) and len(c) >= 3:
            return Polygon(c)
        txt = pm.xpath('.//*[local-name()="LineString"]//*[local-name()="coordinates"]/text()')
        if txt and (c := parse_coords(txt[0])) and len(c) >= 2:
            return LineString(c)
        polys = []
        for nodo in pm.xpath('.//*[local-name()="MultiGeometry"]//*[local-name()="Polygon"]'):
            t = nodo.xpath('.//*[local-name()="outerBoundaryIs"]//*[local-name()="LinearRing"]//*[local-name()="coordinates"]/text()')
            if t and (c := parse_coords(t[0])) and len(c) >= 3:
                polys.append(Polygon(c))
        return MultiPolygon(polys) if polys else None

    def kmz_to_gdf(kmz_path):
        zona = Path(kmz_path).stem
        tmp = tempfile.mkdtemp()
        try:
            with zipfile.ZipFile(kmz_path) as zf:
                zf.extractall(tmp)
            kmls = list(Path(tmp).rglob("*.kml")) + list(Path(tmp).rglob("*.KML"))
            feats = []
            for fp in kmls:
                try:
                    tree = etree.parse(str(fp))
                except Exception:
                    continue
                for pm in tree.xpath('.//*[local-name()="Placemark"]'):
                    g = geom(pm)
                    if g:
                        name = (pm.xpath('.//*[local-name()="name"]/text()') or ["Sin_nombre"])[0]
                        feats.append({"geometry": g, "zona": zona, "subzona": name})
            return gpd.GeoDataFrame(feats, crs="EPSG:4326") if feats else None
        finally:
            shutil.rmtree(tmp, ignore_errors=True)

    frames = [kmz_to_gdf(f) for f in KMZ_FILES]
    frames = [g for g in frames if g is not None and not g.empty]
    return pd.concat(frames, ignore_index=True).to_crs(4326) if frames else \
           gpd.GeoDataFrame(columns=["geometry", "zona", "subzona"], crs="EPSG:4326")

# ----------------------- GEOCODIFICACIÓN ------------------------
def build_address(r):
    comps = [str(r.get(c, "")).strip()
             for c in ["Buyer Address1", "Buyer Address1 Number", "Buyer City"]
             if str(r.get(c, "")).strip() and str(r.get(c, "")).lower() != "nan"]
    return ", ".join(comps) + ", Colombia"

def geocode_enhanced(cli, address, retries=3):
    for _ in range(retries):
        try:
            res = cli.geocode(address, region="co", language="es")
            if res:
                loc = res[0]["geometry"]["location"]
                return loc["lat"], loc["lng"], "completa"
        except Exception:
            time.sleep(0.8)
    try:
        city = address.split(",")[-2] + ", Colombia"
        res = cli.geocode(city, region="co", language="es")
        if res:
            loc = res[0]["geometry"]["location"]
            return loc["lat"], loc["lng"], "ciudad"
    except Exception:
        pass
    return None, None, "fallida"

def geocode_with_cache(cli, address):
    hit = cache_df.loc[cache_df.full_address == address]
    if not hit.empty:
        r = hit.iloc[0]
        return r.latitude, r.longitude, r.geocode_status, True
    lat, lon, stat = geocode_enhanced(cli, address)
    cache_df.loc[len(cache_df)] = [address, lat, lon, stat]
    return lat, lon, stat, False

# ---------------- PRE-CLASIFICACIÓN (estilo ejecutivo) ----------
def crear_preclasificacion(df_geo) -> io.BytesIO:
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    total = len(df_geo)
    base_df = df_geo[["Country", "Tracking Number", "subzona"]].rename(columns={"subzona": "Zona"})
    reg_df  = pd.DataFrame({"Consecutivo": range(1, total + 1),
                            "Tracking Number": ["" for _ in range(total)],
                            "Zona": ["" for _ in range(total)]})

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        base_df.to_excel(writer, sheet_name="Base", index=False)
        reg_df.to_excel(writer, sheet_name="Registro", index=False, startrow=2)

        ws_reg  = writer.sheets["Registro"]
        ws_base = writer.sheets["Base"]

        # --- Fila 1: métricas ---
        for c in ws_reg[1]:
            c.font = c.font.copy(sz=18, bold=True)
        ws_reg.row_dimensions[1].height = 32

        last_row = ws_reg.max_row

        # Fórmulas Zona (col C) y cabeceras dinámicas
        for r in range(4, last_row + 1):
            ws_reg[f"C{r}"] = f'=IFERROR(VLOOKUP($B{r},Base!$B:$C,2,FALSE),"")'
        ws_reg["A1"] = f'=COUNTA(B4:B{last_row})'
        ws_reg["B1"] = f'=IFERROR(LOOKUP(2,1/(B4:B{last_row}<>""),C4:C{last_row}),"")'
        ws_reg["C1"] = f'=IFERROR(LOOKUP(2,1/(B4:B{last_row + 5}<>""),B4:B{last_row + 5}),"")'

        # --- Estilo cabecera de datos (fila 3) ---
        header_fill = PatternFill("solid", fgColor="005BEA")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws_reg[3]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Ajuste de columnas
        for col, w in zip("ABC", (12, 28, 20)):
            ws_reg.column_dimensions[col].width = w

        # --- Congelar fila 1 ---
        ws_reg.freeze_panes = "A2"

        # --- Insertar logo corporativo en D1 ---
        if os.path.exists(LOGO_PATH):
            img = XLImage(LOGO_PATH)
            img.width, img.height = 40, 40
            ws_reg.add_image(img, "E1")

        # --- Diseño minimalista: quitar gridlines y bordes finos ---
        ws_reg.sheet_view.showGridLines = False
        thin = Border(left=Side(style="thin", color="D0D0D0"),
                      right=Side(style="thin", color="D0D0D0"),
                      top=Side(style="thin", color="D0D0D0"),
                      bottom=Side(style="thin", color="D0D0D0"))
        for row in ws_reg.iter_rows(min_row=3, max_row=last_row, min_col=1, max_col=3):
            for cell in row:
                cell.border = thin
                cell.alignment = Alignment(vertical="center")

        # --- Protección: Col A y B editables; Col C bloqueada ---
        for row in ws_reg.iter_rows(min_row=1, max_row=last_row + 5, min_col=1, max_col=3):
            for cell in row:
                lock = cell.column_letter == "C"
                cell.protection = Protection(locked=lock)

        ws_reg.protection.sheet = True
        ws_reg.protection.set_password(PROT_PASS)  # ← contraseña

        # Hoja Base oculta
        ws_base.sheet_state = "veryHidden"

    buf.seek(0)
    return buf

# --------------- PROCESAR UN ARCHIVO SUBIDO ---------------------
def procesar_archivo(uploaded_file):
    df = pd.read_excel(uploaded_file).copy().reset_index(drop=True)
    df["full_address"] = df.apply(build_address, axis=1)

    cli = googlemaps.Client(API_KEY)
    lat, lon, stat = [], [], []
    prog = st.progress(0, text=f"Geocodificando: {uploaded_file.name}")
    for i, addr in enumerate(df.full_address):
        la, lo, stt, hit = geocode_with_cache(cli, addr)
        lat.append(la); lon.append(lo); stat.append(stt)
        prog.progress((i + 1) / len(df))
    prog.empty()
    df["latitude"], df["longitude"], df["geocode_status"] = lat, lon, stat
    cache_df.to_csv(CACHE_FILE, index=False)

    # --- Unión espacial ---
    gdf_pol = load_polygons()
    gdf_pts = gpd.GeoDataFrame(
        df[df.latitude.notna() & df.longitude.notna()].reset_index().rename(columns={'index': 'idx'}),
        geometry=[Point(xy) for xy in zip(df.longitude, df.latitude)], crs=4326
    )
    join = gpd.sjoin(gdf_pts, gdf_pol[["zona", "subzona", "geometry"]],
                     how="left", predicate="within").drop_duplicates(subset='idx').set_index('idx')
    df = df.join(join[["zona", "subzona"]])

    # --- Reglas adicionales ---
    buyer_lower = df["Buyer City"].astype(str).str.lower()
    mask_other  = buyer_lower == "other"
    mask_bq     = buyer_lower == "barranquilla"
    mask_override = (~mask_other) & (~mask_bq)

    df.loc[mask_other, ["zona", "subzona"]] = "Verificar Zona"
    df.loc[mask_override & df["Buyer City"].notna(), "subzona"] = df.loc[mask_override, "Buyer City"]

    df["zona"].fillna("Verificar Zona", inplace=True)
    df["subzona"].fillna("Verificar Zona", inplace=True)

    sin_zona = (df["zona"] == "Verificar Zona").sum()

    # --- Crear archivo final ---
    pre_buf = crear_preclasificacion(df)

    # --- Actualizar hoja master en Google Sheets ---
    try:
        update_master(df)
    except Exception as e:
        st.warning(f"No se pudo actualizar el master en Google Sheets: {e}")

    nombre_final = f"{Path(uploaded_file.name).stem}_zonificado.xlsx"

    return {"filename": nombre_final,
            "buffer": pre_buf,
            "stats": {"total": len(df),
                      "con_zona": len(df) - sin_zona,
                      "sin_zona": sin_zona}}

# --------------------------- INTERFAZ ---------------------------
st.header("🗺️ Geolocalización y Pre-Clasificación")
st.caption("Cargue hasta 5 archivos Excel a la vez.")

uploads = st.file_uploader("⬆️  Seleccione archivos",
                           type=["xlsx", "xls"], accept_multiple_files=True)

if uploads and len(uploads) > 5:
    st.error("Solo se permiten hasta 5 archivos simultáneamente.")

if uploads and st.button("🚀 Procesar Archivos", type="primary"):
    with st.spinner("Procesando archivos…"):
        resultados = [procesar_archivo(u) for u in uploads[:5]]

        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for r in resultados:
                zf.writestr(r["filename"], r["buffer"].getvalue())
        zip_buf.seek(0)

    st.success("¡Procesamiento completo!")

    for r in resultados:
        with st.expander(f"📄 {r['filename']}"):
            st.write(f"Total: {r['stats']['total']}")
            st.write(f"Con zona: {r['stats']['con_zona']}")
            st.write(f"Sin zona: {r['stats']['sin_zona']}")

    st.download_button("📥 Descargar ZIP con archivos procesados",
                       data=zip_buf,
                       file_name="Archivos_Zonificados.zip",
                       mime="application/zip")
elif not uploads:
    st.info("Arrastre o seleccione hasta 5 archivos Excel para habilitar el botón.")